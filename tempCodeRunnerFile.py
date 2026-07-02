"""
WMD Rainfall Intelligence System — Flask Backend
Handles the real WMD Excel format:
  - River Gauge Data (River, Gauge Station, District, Gauge Level, Trend, Danger Level)
  - Rainfall Data (auto-detected)
  - Reservoir Data (auto-detected)
Run:
  pip install -r requirements.txt
  py app.py
  Open → http://127.0.0.1:5000
"""

from flask import Flask, request, jsonify, render_template, session
import pandas as pd
import numpy as np
import io, traceback, json, re
from datetime import date, timedelta
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "wmd-rainfall-secret-2026"
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

# ── Parse the real WMD Excel format ──────────────────
def parse_wmd_excel(content_bytes):
    """
    Reads the complex WMD Excel which has:
    - Multiple sections (RIVER GAUGE DATA, RAINFALL DATA, RESERVOIR DATA)
    - Date in a merged header row like 'DATE:28-07-2025 ...'
    - Headers mixed in rows
    Returns a dict of section dataframes + metadata
    """
    wb = load_workbook(io.BytesIO(content_bytes), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    print(f"[DEBUG] Total rows in file: {len(all_rows)}")
    print(f"[DEBUG] First 5 rows: {all_rows[:5]}")

    # ── Extract date from header ──
    file_date = None
    for row in all_rows[:5]:
        for cell in row:
            if cell and isinstance(cell, str) and 'DATE:' in cell.upper():
                m = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', str(cell))
                if m:
                    try:
                        file_date = pd.to_datetime(m.group(1), dayfirst=True).date()
                    except Exception:
                        pass
                break

    print(f"[DEBUG] Extracted date: {file_date}")

    # ── Find section boundaries ──
    sections = {}
    current_section = None
    header_row_idx  = None
    data_rows = []

    SECTION_KEYWORDS = {
        'RIVER GAUGE':  'river_gauge',
        'RAINFALL':     'rainfall',
        'RESERVOIR':    'reservoir',
        'FLOOD':        'river_gauge',
    }

    def flush_section():
        nonlocal current_section, header_row_idx, data_rows
        if current_section and header_row_idx is not None and data_rows:
            headers = [str(h).replace('\n', ' ').strip() if h else f'Col{i}'
                       for i, h in enumerate(all_rows[header_row_idx])]
            df = pd.DataFrame(data_rows, columns=headers[:len(data_rows[0])])
            df = df.dropna(how='all')
            sections[current_section] = df
            print(f"[DEBUG] Section '{current_section}': {df.shape}, cols={headers}")
        current_section = None
        header_row_idx  = None
        data_rows       = []

    for i, row in enumerate(all_rows):
        row_text = ' '.join(str(c).upper() for c in row if c)

        # Detect new section
        for kw, sec_name in SECTION_KEYWORDS.items():
            if kw in row_text and header_row_idx is None:
                flush_section()
                current_section = sec_name
                break

        # Look for the actual column header row (has multiple non-None cells)
        if current_section and header_row_idx is None:
            non_null = sum(1 for c in row if c)
            if non_null >= 3 and i > 0:
                # Check it looks like a header (contains words like 'Name', 'District', 'Level', 'Station', 'No')
                row_str = row_text
                if any(w in row_str for w in ['NAME', 'DISTRICT', 'LEVEL', 'STATION', 'NO.', 'SL', 'RAINFALL', 'DATE']):
                    header_row_idx = i
            continue

        # Collect data rows
        if current_section and header_row_idx is not None:
            non_null = sum(1 for c in row if c)
            if non_null >= 1:
                data_rows.append(list(row))

    flush_section()

    # ── If no sections detected, try reading as plain table ──
    if not sections:
        print("[DEBUG] No sections found, trying plain pandas read")
        df = pd.read_excel(io.BytesIO(content_bytes), header=None)
        # Find first numeric row
        for i, row in df.iterrows():
            nums = row.apply(lambda x: pd.to_numeric(x, errors='coerce')).notna().sum()
            if nums >= 2:
                header = df.iloc[i-1] if i > 0 else pd.RangeIndex(len(df.columns))
                data   = df.iloc[i:]
                data.columns = [str(h) for h in header]
                sections['data'] = data
                break

    return {
        'sections':  sections,
        'file_date': str(file_date) if file_date else None,
        'raw_rows':  len(all_rows),
    }


def analyse_wmd(parsed):
    """Convert parsed WMD sections into the standard result JSON."""
    sections  = parsed['sections']
    file_date = parsed['file_date']
    result    = {
        'file_date': file_date,
        'sections_found': list(sections.keys()),
    }

    # ── RIVER GAUGE section ──
    river_data = []
    alerts     = []
    if 'river_gauge' in sections:
        df = sections['river_gauge'].copy()
        print(f"[DEBUG] River gauge cols: {df.columns.tolist()}")
        print(f"[DEBUG] River gauge head:\n{df.head(5)}")

        # Find relevant columns by keyword
        river_col   = next((c for c in df.columns if 'river' in c.lower()), None)
        station_col = next((c for c in df.columns if 'station' in c.lower() or 'gauge' in c.lower() and 'level' not in c.lower()), None)
        district_col= next((c for c in df.columns if 'district' in c.lower()), None)
        level_col   = next((c for c in df.columns if 'level' in c.lower() and 'danger' not in c.lower() and 'extreme' not in c.lower()), None)
        trend_col   = next((c for c in df.columns if 'trend' in c.lower()), None)
        danger_col  = next((c for c in df.columns if 'danger' in c.lower() and 'extreme' not in c.lower()), None)
        ext_danger_col = next((c for c in df.columns if 'extreme' in c.lower()), None)

        print(f"[DEBUG] Mapped cols: river={river_col}, station={station_col}, level={level_col}, danger={danger_col}")

        for _, row in df.iterrows():
            try:
                river   = str(row[river_col]).strip()   if river_col   else ''
                station = str(row[station_col]).strip() if station_col else ''
                district= str(row[district_col]).strip() if district_col else ''
                level   = pd.to_numeric(row[level_col],    errors='coerce') if level_col    else None
                danger  = pd.to_numeric(row[danger_col],   errors='coerce') if danger_col   else None
                ext_dng = pd.to_numeric(row[ext_danger_col], errors='coerce') if ext_danger_col else None
                trend   = str(row[trend_col]).strip() if trend_col and row[trend_col] else ''

                if river in ('', 'nan', 'None') and station in ('', 'nan', 'None'):
                    continue

                status = 'Normal'
                if level is not None and not np.isnan(level):
                    if ext_dng is not None and not np.isnan(ext_dng) and level >= ext_dng:
                        status = 'EXTREME DANGER'
                    elif danger is not None and not np.isnan(danger) and level >= danger:
                        status = 'DANGER'
                    elif danger is not None and not np.isnan(danger) and level >= danger * 0.9:
                        status = 'WARNING'

                entry = {
                    'river':    river,
                    'station':  station,
                    'district': district,
                    'level':    round(float(level), 2) if level is not None and not np.isnan(level) else None,
                    'trend':    trend,
                    'danger':   round(float(danger), 2) if danger is not None and not np.isnan(danger) else None,
                    'ext_danger': round(float(ext_dng), 2) if ext_dng is not None and not np.isnan(ext_dng) else None,
                    'status':   status,
                }
                river_data.append(entry)

                if status in ('DANGER', 'EXTREME DANGER'):
                    alerts.append({
                        'river':   river,
                        'station': station,
                        'level':   entry['level'],
                        'danger':  entry['danger'],
                        'status':  status,
                        'date':    file_date or str(date.today()),
                    })

            except Exception as e:
                print(f"[DEBUG] Row parse error: {e}")
                continue

    # ── RAINFALL section ──
    rainfall_summary = {}
    total_rainfall   = 0.0
    if 'rainfall' in sections:
        df = sections['rainfall'].copy()
        nums = df.select_dtypes(include=[np.number])
        if not nums.empty:
            total_rainfall = round(float(nums.values[~np.isnan(nums.values)].sum()), 2)
            # Per-district or per-station
            dist_col = next((c for c in df.columns if 'district' in c.lower() or 'station' in c.lower()), None)
            rain_col = next((c for c in df.columns if 'rain' in c.lower() or 'mm' in c.lower()), nums.columns[0] if len(nums.columns) else None)
            if dist_col and rain_col:
                for _, row in df.iterrows():
                    d = str(row[dist_col]).strip()
                    r = pd.to_numeric(row[rain_col], errors='coerce')
                    if d not in ('', 'nan', 'None') and not np.isnan(r):
                        rainfall_summary[d] = round(float(r), 2)

    # ── River stats ──
    levels_with_data = [r for r in river_data if r['level'] is not None]
    river_stats = {
        'total_stations':   len(river_data),
        'stations_with_data': len(levels_with_data),
        'danger_count':     len([r for r in river_data if r['status'] == 'DANGER']),
        'extreme_count':    len([r for r in river_data if r['status'] == 'EXTREME DANGER']),
        'warning_count':    len([r for r in river_data if r['status'] == 'WARNING']),
        'normal_count':     len([r for r in river_data if r['status'] == 'Normal']),
        'max_level':        max((r['level'] for r in levels_with_data), default=0),
        'falling_count':    len([r for r in river_data if 'fall' in r['trend'].lower()]),
        'rising_count':     len([r for r in river_data if 'ris' in r['trend'].lower()]),
        'steady_count':     len([r for r in river_data if 'steady' in r['trend'].lower()]),
    }

    # ── Chart data: gauge levels by station ──
    chart_stations = [r['station'] or r['river'] for r in levels_with_data]
    chart_levels   = [r['level'] for r in levels_with_data]
    chart_danger   = [r['danger'] for r in levels_with_data]

    return {
        'file_date':        file_date,
        'river_data':       river_data,
        'river_stats':      river_stats,
        'alerts':           alerts,
        'rainfall_summary': rainfall_summary,
        'total_rainfall':   total_rainfall,
        'chart': {
            'stations': chart_stations,
            'levels':   chart_levels,
            'danger':   chart_danger,
        },
        'sections_found': list(sections.keys()),
    }


# ── Q&A ───────────────────────────────────────────
def answer_question(question, data):
    q = question.lower().strip()
    rs   = data.get('river_stats', {})
    rd   = data.get('river_data',  [])
    alts = data.get('alerts',      [])
    fd   = data.get('file_date',   'N/A')

    if any(w in q for w in ['danger', 'alert', 'flood', 'warning', 'risk', 'breach']):
        dc = rs.get('danger_count', 0)
        ec = rs.get('extreme_count', 0)
        if alts:
            lines = ''.join([
                f"<br>⚠️ <strong>{a['river']} at {a['station']}</strong> — "
                f"Level: {a['level']} m (Danger: {a['danger']} m) [{a['status']}]"
                for a in alts
            ])
            return (f"🚨 <strong>{dc} stations at DANGER</strong>, "
                    f"{ec} at EXTREME DANGER as of {fd}:{lines}")
        return f"✅ No danger-level stations detected as of {fd}. All rivers are within normal limits."

    if any(w in q for w in ['today', 'current', 'now', 'date', 'report']):
        return (f"📅 This report is for date: <strong>{fd}</strong><br>"
                f"• Stations monitored: {rs.get('total_stations', 0)}<br>"
                f"• Stations with gauge data: {rs.get('stations_with_data', 0)}<br>"
                f"• Danger: {rs.get('danger_count', 0)} | Warning: {rs.get('warning_count', 0)} | Normal: {rs.get('normal_count', 0)}")

    if any(w in q for w in ['trend', 'rising', 'falling', 'steady']):
        return (f"📈 River trends as of {fd}:<br>"
                f"• 🔴 Rising: <strong>{rs.get('rising_count', 0)}</strong> stations<br>"
                f"• 🟡 Steady: <strong>{rs.get('steady_count', 0)}</strong> stations<br>"
                f"• 🟢 Falling: <strong>{rs.get('falling_count', 0)}</strong> stations")

    if any(w in q for w in ['rainfall', 'rain', 'mm', 'precipitation']):
        tr = data.get('total_rainfall', 0)
        rs_data = data.get('rainfall_summary', {})
        if rs_data:
            lines = ' | '.join([f"{k}: {v} mm" for k, v in list(rs_data.items())[:5]])
            return f"🌧️ Total rainfall: <strong>{tr} mm</strong><br>{lines}"
        return f"🌧️ Total rainfall recorded: <strong>{tr} mm</strong>. No district-wise breakdown available."

    if any(w in q for w in ['river', 'station', 'gauge', 'level']):
        if rd:
            with_data = [r for r in rd if r['level'] is not None][:5]
            lines = ''.join([
                f"<br>• <strong>{r['river']}</strong> at {r['station']}: "
                f"{r['level']} m ({r['trend']}) [{r['status']}]"
                for r in with_data
            ])
            return f"🌊 River gauge levels as of {fd}:{lines}"
        return "No river gauge data found in the uploaded file."

    if any(w in q for w in ['max', 'maximum', 'highest']):
        m = rs.get('max_level', 0)
        top = [r for r in rd if r['level'] == m]
        if top:
            t = top[0]
            return (f"📈 Highest gauge level: <strong>{m} m</strong> "
                    f"at <strong>{t['river']} — {t['station']}</strong> [{t['status']}]")
        return f"📈 Maximum gauge level recorded: <strong>{m} m</strong>"

    if any(w in q for w in ['how many', 'count', 'total station', 'number']):
        return (f"🔢 Total stations monitored: <strong>{rs.get('total_stations', 0)}</strong><br>"
                f"• With gauge data: {rs.get('stations_with_data', 0)}<br>"
                f"• Danger/Extreme: {rs.get('danger_count', 0) + rs.get('extreme_count', 0)}<br>"
                f"• Normal: {rs.get('normal_count', 0)}")

    # Default full summary
    return (f"📋 <strong>WMD Daily Flood Report — {fd}</strong><br>"
            f"• Stations monitored: {rs.get('total_stations', 0)} "
            f"({rs.get('stations_with_data', 0)} with data)<br>"
            f"• 🔴 Danger: {rs.get('danger_count', 0)} | "
            f"⚫ Extreme: {rs.get('extreme_count', 0)} | "
            f"🟡 Warning: {rs.get('warning_count', 0)} | "
            f"🟢 Normal: {rs.get('normal_count', 0)}<br>"
            f"• Trends — Rising: {rs.get('rising_count', 0)} | "
            f"Steady: {rs.get('steady_count', 0)} | "
            f"Falling: {rs.get('falling_count', 0)}<br>"
            f"• Total Rainfall: {data.get('total_rainfall', 0)} mm<br>"
            f"<em>Ask: danger alerts · river levels · trends · rainfall · how many stations</em>")


# ── Routes ────────────────────────────────────────
@app.route('/')
def welcome():
    return render_template('welcome.html')

@app.route('/portal')
def portal():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file in request.'}), 400
        f = request.files['file']
        if not f or not f.filename:
            return jsonify({'error': 'No file selected.'}), 400
        fname = f.filename.lower()
        if not any(fname.endswith(e) for e in ['.xlsx', '.xls', '.csv']):
            return jsonify({'error': 'Only .xlsx / .xls / .csv supported.'}), 400

        content = f.read()
        if not content:
            return jsonify({'error': 'File is empty.'}), 400

        print(f'[DEBUG] File: {f.filename} ({len(content)} bytes)')

        if fname.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
            # simple CSV path
            result = {'filename': f.filename, 'mode': 'csv', 'river_data': [],
                      'river_stats': {}, 'alerts': [], 'rainfall_summary': {},
                      'total_rainfall': 0, 'chart': {'stations':[],'levels':[],'danger':[]},
                      'file_date': str(date.today()), 'sections_found': ['data']}
        else:
            parsed = parse_wmd_excel(content)
            result = analyse_wmd(parsed)
            result['filename'] = f.filename
            result['mode'] = 'wmd_excel'

        session['data_cache'] = json.dumps(result)
        print(f'[DEBUG] Done — stations: {len(result.get("river_data", []))}, alerts: {len(result.get("alerts", []))}')
        return jsonify(result)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/ask', methods=['POST'])
def ask():
    try:
        body = request.get_json(silent=True) or {}
        question = (body.get('question') or '').strip()
        if not question:
            return jsonify({'answer': 'Please type a question.'})
        raw = session.get('data_cache')
        if not raw:
            return jsonify({'answer': '⚠️ Please upload your WMD Excel file first.'})
        data = json.loads(raw)
        return jsonify({'answer': answer_question(question, data)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'answer': f'Error: {e}'}), 500


if __name__ == '__main__':
    print('=' * 50)
    print('🌧️  WMD Rainfall Intelligence System')
    print('=' * 50)
    app.run(debug=True, port=5000)